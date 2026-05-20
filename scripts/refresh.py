#!/usr/bin/env python3
"""
Compliance Watch — daily refresh script.

Reads the newest CSV file from raw/ (Origami's scheduled export),
joins against scripts/property_reference.csv to get RVP/region mapping,
computes summary metrics, and writes data/dashboard_data.json.

Run by the GitHub Action whenever a new CSV is committed to raw/.

This is the V1 refresh: builds the basic data file. Future versions will add:
  - Archive auto-population (move resolved permits to archive)
  - Daily history snapshot (for real month-over-month trends)
  - Note preservation across refreshes (notes live in Azure Table Storage,
    not in this JSON — the dashboard pulls them separately at runtime)
"""
import csv
import json
import sys
import re
from datetime import date, datetime, timedelta
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
RAW_DIR = REPO / 'raw'
DATA_DIR = REPO / 'data'
SCRIPTS_DIR = REPO / 'scripts'

REFERENCE_CSV = SCRIPTS_DIR / 'property_reference.csv'
OUTPUT_JSON = DATA_DIR / 'dashboard_data.json'

# Window thresholds (match the Origami report's filter)
EXPIRING_SOON_DAYS = 30


def newest_raw_file() -> Path:
    """Find the newest CSV or XLSX in raw/, by file mtime."""
    candidates = list(RAW_DIR.glob('*.csv')) + list(RAW_DIR.glob('*.xlsx'))
    candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        sys.exit(f"ERROR: no CSV or XLSX files found in {RAW_DIR}")
    return candidates[0]


def newest_raw_csv() -> Path:
    """Backward-compatible alias."""
    return newest_raw_file()


def read_rows(raw_path: Path):
    """Yield row dicts from either a CSV or an XLSX file.

    For XLSX, expects Origami's scheduled-report format: headers on row 5
    (after 4 metadata rows). Falls back to standard CSV reading otherwise.
    """
    suffix = raw_path.suffix.lower()
    if suffix == '.csv':
        with raw_path.open(encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader, start=2):
                yield i, row
        return

    if suffix == '.xlsx':
        # Lazy import so the dependency is only required when XLSX is used.
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("ERROR: openpyxl required to read XLSX. "
                     "Add 'openpyxl' to your GitHub Action's pip install step.")
        wb = load_workbook(raw_path, read_only=True, data_only=True)
        ws = wb.active
        # Find the header row: the first row containing the string 'Title'.
        # Origami's scheduled report puts headers on row 5; manual exports
        # may differ. Scan the first 10 rows to be safe.
        header = None
        header_row = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            row_strs = [str(c) if c is not None else '' for c in row]
            if 'Title' in row_strs:
                header = row_strs
                header_row = r_idx
                break
        if header is None:
            sys.exit(f"ERROR: could not find header row in {raw_path.name}")

        # Iterate data rows
        for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 2):
            if all(c is None or str(c).strip() == '' for c in row):
                continue  # skip blank rows
            row_dict = {}
            for col_idx, value in enumerate(row):
                if col_idx >= len(header):
                    break
                col_name = header[col_idx]
                # Convert datetime objects to ISO date strings for the parser
                if hasattr(value, 'isoformat'):
                    row_dict[col_name] = value.strftime('%Y-%m-%d')
                else:
                    row_dict[col_name] = '' if value is None else str(value)
            yield r_idx, row_dict
        return

    sys.exit(f"ERROR: unsupported file type {suffix}. Expected .csv or .xlsx.")


def load_property_reference():
    """Returns (ref_by_code, name_to_code).

    ref_by_code: dict keyed by property code -> {display_name, rvp, region}
    name_to_code: dict keyed by uppercase Origami name -> property code
                  (used to look up which property a CSV row belongs to)
    """
    if not REFERENCE_CSV.exists():
        sys.exit(f"ERROR: property reference not found at {REFERENCE_CSV}")
    ref_by_code = {}
    name_to_code = {}
    with REFERENCE_CSV.open() as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row['Property Code'].strip().upper()
            display_name = row['Display Name'].strip()
            origami_name = row.get('Origami Name', '').strip() or display_name
            ref_by_code[code] = {
                'display_name': display_name,
                'rvp': row['RVP/SVP'].strip(),
                'region': row['Region'].strip(),
            }
            # Build case-insensitive lookup
            name_to_code[origami_name.upper().strip()] = code
            # Also accept the display name as a fallback match
            name_to_code[display_name.upper().strip()] = code
    return ref_by_code, name_to_code


def parse_date(s: str):
    """Origami exports MM/DD/YYYY. Tolerate ISO and blanks."""
    if not s or not s.strip():
        return None
    s = s.strip()
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def extract_property_code(row, name_to_code):
    """Find the property code for a row.

    Strategy:
    1. Origami's scheduled report puts just the property name in 'Location'
       (e.g., 'SAFE HARBOR ALLEN HARBOR'). We match this against the
       Origami Name column in property_reference.csv (case-insensitive).
    2. If the Location is in 'CODE - Name' format (manual export), we
       still parse the code from the prefix.
    3. Fallback to other column names in case Origami's format changes.
    """
    loc = (row.get('Location') or '').strip()

    # First: case-insensitive name lookup (the scheduled report path)
    if loc and loc.upper() in name_to_code:
        return name_to_code[loc.upper()]

    # Second: 'CODE - Name' format (manual export path)
    if loc and ' - ' in loc:
        code_candidate = loc.split(' - ', 1)[0].strip().upper()
        # Validate it looks like a property code (short alphanumeric)
        if code_candidate and len(code_candidate) <= 6 and code_candidate.isalnum():
            return code_candidate

    # Third: explicit Property Code column (in case Origami's format changes)
    for key in ('Property Code', 'PropertyCode', 'Property', 'Location Code'):
        if row.get(key):
            return row[key].strip().upper()

    return None


def extract_title(row):
    """Get the permit title from Origami's 'Title' column (or fallbacks)."""
    for key in ('Title', 'Document Title', 'Permit Title'):
        if row.get(key):
            return row[key].strip()
    return ''


def process_csv(csv_path: Path, ref: dict, name_to_code: dict, today: date):
    """Read the CSV (or XLSX), filter to relevant permits, build records."""
    horizon = today + timedelta(days=EXPIRING_SOON_DAYS)
    active = []
    skipped_unknown_property = []

    for i, row in read_rows(csv_path):
        code = extract_property_code(row, name_to_code)
        if not code:
            continue
        if code not in ref:
            skipped_unknown_property.append((i, code))
            continue

        exp = parse_date(row.get('Expiration Date') or row.get('Expiration'))
        ren = parse_date(row.get('Renewal Start Date') or row.get('Renewal Start'))
        if not exp:
            continue

        # Apply the same filter as the Origami saved report:
        # expiration <= today + 30 days AND renewal start <= today + 30 days
        if exp > horizon:
            continue
        if ren and ren > horizon:
            continue

        days_to_exp = (exp - today).days
        status = 'expired' if days_to_exp < 0 else 'expiring_soon'

        title = extract_title(row)
        # Origami's export doesn't include a separate document ID, so we
        # generate a stable per-row identifier from property code + title hash.
        # In a future refresh we can pull a proper ID if Origami's format expands.
        import hashlib
        tid = hashlib.md5(f"{code}|{title}|{exp.isoformat()}".encode()).hexdigest()[:8]
        permit_id = f"{code}-{tid}"

        info = ref[code]
        active.append({
            'permit_id': permit_id,
            'title': title,
            'property_code': code,
            'display_name': info['display_name'],
            'rvp': info['rvp'],
            'region': info['region'],
            'renewal_start': ren.isoformat() if ren else None,
            'expiration': exp.isoformat(),
            'status': status,
            'days_to_exp': days_to_exp,
        })

    if skipped_unknown_property:
        print(f"WARNING: skipped {len(skipped_unknown_property)} rows with unknown property codes:")
        for line, code in skipped_unknown_property[:10]:
            print(f"  line {line}: {code}")
        if len(skipped_unknown_property) > 10:
            print(f"  ... and {len(skipped_unknown_property) - 10} more")

    return active


def total_permits_by_property(csv_path: Path, ref: dict, name_to_code: dict):
    """Count total permits per property (all rows, not just expired/expiring)."""
    totals = {}
    for _, row in read_rows(csv_path):
        code = extract_property_code(row, name_to_code)
        if code and code in ref:
            totals[code] = totals.get(code, 0) + 1
    return totals


def build_summary(active, ref, totals, today):
    total_permits = sum(totals.values())
    total_expired = sum(1 for p in active if p['status'] == 'expired')
    total_expiring = sum(1 for p in active if p['status'] == 'expiring_soon')
    properties_with_issues = len({p['property_code'] for p in active})
    compliance = round(100 * (total_permits - total_expired) / total_permits, 1) if total_permits else 100.0
    return {
        'as_of': today.isoformat(),
        'total_permits': total_permits,
        'total_properties': len(ref),
        'total_expired': total_expired,
        'total_expiring': total_expiring,
        'properties_with_issues': properties_with_issues,
        'overall_compliance': compliance,
        'total_regions': len({info['region'] for info in ref.values()}),
    }


def build_region_summary(active, ref, totals):
    # Group properties by region for totals
    by_region = {}
    for code, info in ref.items():
        r = info['region']
        bucket = by_region.setdefault(r, {
            'region': r, 'rvp': info['rvp'],
            'total_permits': 0, 'expired': 0, 'expiring_soon': 0,
            'property_count': 0, 'properties_with_issues': set(),
        })
        bucket['total_permits'] += totals.get(code, 0)
        bucket['property_count'] += 1
    for p in active:
        b = by_region[p['region']]
        if p['status'] == 'expired':
            b['expired'] += 1
        else:
            b['expiring_soon'] += 1
        b['properties_with_issues'].add(p['property_code'])

    out = []
    for r, b in by_region.items():
        total = b['total_permits']
        compliance = round(100 * (total - b['expired']) / total, 1) if total else 100.0
        out.append({
            'region': b['region'], 'rvp': b['rvp'],
            'total_permits': total,
            'expired': b['expired'], 'expiring_soon': b['expiring_soon'],
            'pct_expired': round(100 * b['expired'] / total, 1) if total else 0,
            'pct_expiring_soon': round(100 * b['expiring_soon'] / total, 1) if total else 0,
            'compliance_score': compliance,
            'property_count': b['property_count'],
            'properties_with_issues': len(b['properties_with_issues']),
        })
    out.sort(key=lambda x: (x['compliance_score'], -x['expired']))
    return out


def build_property_summary(active, ref, totals):
    by_property = {}
    for code, info in ref.items():
        by_property[code] = {
            'code': code,
            'display_name': info['display_name'],
            'rvp': info['rvp'],
            'region': info['region'],
            'total_permits': totals.get(code, 0),
            'expired': 0,
            'expiring_soon': 0,
        }
    for p in active:
        b = by_property[p['property_code']]
        if p['status'] == 'expired':
            b['expired'] += 1
        else:
            b['expiring_soon'] += 1

    out = []
    for b in by_property.values():
        total = b['total_permits']
        b['compliance_score'] = round(100 * (total - b['expired']) / total, 1) if total else 100.0
        out.append(b)
    out.sort(key=lambda x: (x['compliance_score'], -x['expired']))
    return out


def build_properties(ref, totals):
    out = []
    for code, info in ref.items():
        out.append({
            'code': code,
            'display_name': info['display_name'],
            'rvp': info['rvp'],
            'region': info['region'],
            'total_permits': totals.get(code, 0),
        })
    out.sort(key=lambda x: x['display_name'])
    return out


def load_existing_archive():
    """Preserve the existing archive on each refresh.
    Archive auto-population (moving permits that fell off the active list)
    will be added in a future refresh-script version."""
    if not OUTPUT_JSON.exists():
        return []
    try:
        existing = json.loads(OUTPUT_JSON.read_text())
        return existing.get('archive_permits', [])
    except (json.JSONDecodeError, KeyError):
        return []


def main():
    today = date.today()
    csv_path = newest_raw_csv()
    print(f"Processing: {csv_path.name} (as of {today.isoformat()})")

    ref, name_to_code = load_property_reference()
    print(f"Loaded property reference: {len(ref)} properties")

    totals = total_permits_by_property(csv_path, ref, name_to_code)
    active = process_csv(csv_path, ref, name_to_code, today)
    print(f"Active expired/expiring permits: {len(active)}")

    summary = build_summary(active, ref, totals, today)
    region_summary = build_region_summary(active, ref, totals)
    property_summary = build_property_summary(active, ref, totals)
    properties = build_properties(ref, totals)
    archive = load_existing_archive()

    output = {
        'summary': summary,
        'active_permits': active,
        'region_summary': region_summary,
        'property_summary': property_summary,
        'properties': properties,
        'archive_permits': archive,
    }

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(output, indent=2))
    print(f"\nWrote {OUTPUT_JSON}")
    print(f"  Compliance: {summary['overall_compliance']}%")
    print(f"  Expired: {summary['total_expired']}, Expiring: {summary['total_expiring']}")


if __name__ == '__main__':
    main()
